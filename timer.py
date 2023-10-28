
import tkinter as tk
from tkinter import simpledialog, colorchooser, font
import tkinter.font as tkfont
from tkinter import ttk
import time
import openpyxl
from openpyxl.styles import PatternFill

is_paused = False
paused_time = 0
pause_start_time = 0

TIMER_INTERVALS = {
    "Prepared Speech": {
        "white": 300,  # 5 minutes
        "green": 360,  # 6 minutes
        "yellow": 420,  # 7 minutes
        "red": float('inf')  # Above 7 minutes
    },
    "Table Topics": {
        "white": 60,  # 1 minute
        "green": 90,  # 1 minute 30 seconds
        "yellow": 120,  # 2 minutes
        "red": float('inf')  # No time limit
    },
    "Ice Breaker": {
        "white": 240,  # 4 minutes
        "green": 300,  # 5 minutes
        "yellow": 360,  # 6 minutes
        "red": float('inf')  # No time limit
    },
    "Custom Timer": {}
}

current_option = "Prepared Speech"

BACKGROUND_COLOR = 'white'
FONT_SIZE = 60
SKIP_SECONDS = 5

def update_background():
    if not is_paused:
        elapsed_time = time.time() - start_time - paused_time
        timer_intervals = TIMER_INTERVALS[current_option]

        if current_option == "Custom Timer":
            for color, interval in timer_intervals.items():
                if elapsed_time < interval:
                    background_color = color
                    break
        else:
            if elapsed_time < timer_intervals["white"]:
                background_color = "white"
            elif elapsed_time < timer_intervals["green"]:
                background_color = "#00FF00"
            elif elapsed_time < timer_intervals["yellow"]:
                background_color = "yellow"
            else:
                background_color = "red"

        BACKGROUND_COLOR = background_color
        root.configure(background=BACKGROUND_COLOR)
        clock_label.configure(background=BACKGROUND_COLOR)

    root.after(1000, update_background)

def update_clock():
    global FONT_SIZE
    if not is_paused:
        elapsed_time = time.time() - start_time - paused_time

        if current_option == "Countdown":
            timer_intervals = TIMER_INTERVALS[current_option]
            remaining_time = timer_intervals["white"] - elapsed_time
            if remaining_time < 0:
                remaining_time = 0

            minutes = int(remaining_time // 60)
            seconds = int(remaining_time % 60)
            clock_label.config(text=f"{seconds:02d}", font=("Arial", FONT_SIZE))

            if remaining_time <= 0:
                # Countdown finished, stop the stopwatch
                pause_stopwatch()
                reset_stopwatch()
        else:
            minutes = int(elapsed_time // 60)
            seconds = int(elapsed_time % 60)
            clock_label.config(text=f"{minutes:02d}:{seconds:02d}", font=("Arial", FONT_SIZE))

    clock_label.after(1000, update_clock)

def show_hide_entries():
    if menu.winfo_ismapped():
        menu.pack_forget()
        entry_frame.pack(side=tk.TOP, padx=10, pady=10)
        toolbar.pack_forget()
    else:
        menu.pack(side=tk.TOP, padx=10)
        entry_frame.pack_forget()
        toolbar.pack(side=tk.TOP, padx=10, pady=10)

def start_stopwatch():
    global start_time
    global is_paused
    global paused_time

    if is_paused:
        is_paused = False
        resume_time = time.time()
        paused_time += resume_time - pause_start_time
        start_time = resume_time - paused_time
        update_background()
        update_clock()
        stop_button.config(text="Stop", state=tk.NORMAL)
        reset_button.config(state=tk.NORMAL)
        menu.pack_forget()  # Hide the menuframe

    else:
        start_time = time.time()
        paused_time = 0
        update_background()
        update_clock()
        stop_button.config(text="Stop", state=tk.NORMAL)
        reset_button.config(state=tk.NORMAL)
        menu.pack_forget()  # Hide the menu frame
        toolbar.pack_forget()

def pause_stopwatch():
    global is_paused
    global pause_start_time
    global paused_time

    if is_paused:
        # Already paused, so resume stopwatch
        is_paused = False
        resume_time = time.time()
        paused_time += resume_time - pause_start_time
        update_background()
        update_clock()
        stop_button.config(text="Stop")

    else:
        # Pause the stopwatch
        is_paused = True
        pause_start_time = time.time()
        stop_button.config(text="Resume")

def reset_stopwatch():
    global is_paused
    global start_time
    global paused_time

    is_paused = True
    paused_time = 0
    clock_label.config(text="00:00")
    stop_button.config(state=tk.DISABLED)
    root.configure(background="white")  # Change background to white
    clock_label.configure(background='white')
    reset_button.config(state=tk.DISABLED)
    menu.pack()  # Show the menu frame

def set_option(option):
    global current_option
    current_option = option

    if option == "Custom Timer":
        set_custom_timer_intervals()
    elif option == "Countdown":
        set_countdown_time()

def set_custom_timer_intervals():
    global TIMER_INTERVALS

    # Clear previous custom timer intervals
    TIMER_INTERVALS["Custom Timer"] = {}

    # Get custom timer intervals from user input
    for color in ["white", "green", "yellow", "red"]:
        interval = simpledialog.askinteger(
            "Custom Timer",
            f"Enter the time in seconds for {color} background color:"
        )
        if interval is not None:
            TIMER_INTERVALS["Custom Timer"][color] = interval
        else:
            break

def set_countdown_time():
    global TIMER_INTERVALS

    countdown_time = simpledialog.askinteger(
        "Countdown Timer",
        "Enter the countdown time in seconds:"
    )

    if countdown_time is not None:
        TIMER_INTERVALS["Countdown"] = {
            BACKGROUND_COLOR: countdown_time,
            "green": float('inf'),  # No time limit
            "yellow": float('inf'),  # No time limit
            "red": float('inf')  # No time limit
        }
        root.configure(background=BACKGROUND_COLOR)

last_tap_time = 0
def skip_seconds(event):
    global start_time
    global paused_time
    global is_paused
    global last_tap_time

    skip_time = 5  # Number of seconds to skip

    if is_paused:
        return

    current_time = time.time()
    if current_time - last_tap_time < 0.5:
        # Double tap detected, fast forward by 5 seconds
        new_time = start_time + paused_time + skip_time
        paused_time += new_time - (start_time + paused_time)
        update_clock()
    else:
        last_tap_time = current_time

def open_settings(event):
    global BACKGROUND_COLOR
    global FONT_SIZE

    # Open settings dialog
    settings_dialog = tk.Toplevel(root)

    def change_background_color():
        color = colorchooser.askcolor(title="Select Background Color")
        if color[1] is not None:
            BACKGROUND_COLOR = color[1]
            settings_dialog.configure(background=BACKGROUND_COLOR)
            root.configure(background=BACKGROUND_COLOR)

    def change_font_size():
        global FONT_SIZE
        size = simpledialog.askinteger("Change Font Size", "Enter the font size:")
        if size is not None:
            FONT_SIZE = size
            clock_label.config(font=("Arial", FONT_SIZE))

    def change_font_style():
        global FONT_SIZE
        new_font = tkfont.askfont()
        if new_font is not None:
            FONT_SIZE = new_font.actual("size")
            clock_label.config(font=(new_font.actual("family"), FONT_SIZE))

    settings_dialog.configure(background=BACKGROUND_COLOR)

    background_button = tk.Button(settings_dialog, text="Change Background Color", command=change_background_color)
    background_button.pack(padx=10, pady=10)

    font_button = tk.Button(settings_dialog, text="Change Font Size", command=change_font_size)
    font_button.pack(padx=10, pady=10)

    style_button = tk.Button(settings_dialog, text="Change Font Style", command=change_font_style)
    style_button.pack(padx=10, pady=10)

    settings_dialog.transient(root)
    settings_dialog.grab_set()
    root.wait_window(settings_dialog)

def save_data():
    global members_data
    date = date_entry.get()
    name = name_entry.get()
    speech_type = selected_option.get()
    time_spoken = clock_label.cget("text")

    members_data.append((date, name, speech_type, time_spoken))
    show_hide_entries()

def generate_report():
    report_file = "Toastmasters_Report.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Speech Data"

    sheet.cell(row=1, column=1, value="Date")
    sheet.cell(row=1, column=2, value="Member Name")
    sheet.cell(row=1, column=3, value="Speech Type")
    sheet.cell(row=1, column=4, value="Time Spoken")

    for index, data in enumerate(members_data, start=2):
        date, name, speech_type, time_spoken = data
        sheet.cell(row=index, column=1, value=date)
        sheet.cell(row=index, column=2, value=name)
        sheet.cell(row=index, column=3, value=speech_type)
        sheet.cell(row=index, column=4, value=time_spoken)

    # Apply conditional formatting to highlight qualified and overshoot speeches
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if ":" in cell.value:
                minutes, seconds = map(int, cell.value.split(":"))
                total_seconds = minutes * 60 + seconds
                if current_option == "Prepared Speech" and total_seconds >= 420:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif current_option == "Table Topics" and total_seconds >= 60:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif current_option == "Ice Breaker" and total_seconds >= 360:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    workbook.save(report_file)
    report_label.config(text=f"Report generated and saved as {report_file}")

# Window Control
root = tk.Tk()
root.attributes('-fullscreen', True)
root.protocol('WM_DELETE_WINDOW', root.destroy)
root.configure(background=BACKGROUND_COLOR)
root.bind("<Double-Button-1>", skip_seconds)
root.bind("<Button-3>", open_settings)

# Stopwatch controls
stopwatch_controls_frame = tk.Frame(root)
stopwatch_controls_frame.pack(side=tk.BOTTOM, anchor='center')

start_button = tk.Button(stopwatch_controls_frame, text="Start", command=start_stopwatch)
start_button.pack(side=tk.LEFT, padx=10, pady=10)

stop_button = tk.Button(stopwatch_controls_frame, text="Stop", command=pause_stopwatch, state=tk.DISABLED)
stop_button.pack(side=tk.LEFT, padx=10, pady=10)

reset_button = tk.Button(stopwatch_controls_frame, text="Reset", command=reset_stopwatch, state=tk.DISABLED)
reset_button.pack(side=tk.LEFT, padx=10, pady=10)

generate_report_button = tk.Button(stopwatch_controls_frame, text="Generate Report", command=generate_report)
generate_report_button.pack(side=tk.LEFT, padx=10, pady=10)

# Digital clock
clock_frame = tk.Frame(root)
clock_frame.pack(side=tk.TOP, pady=50)

clock_label = tk.Label(clock_frame, text="00:00", font=("Arial", FONT_SIZE), bg=BACKGROUND_COLOR, fg='black')
clock_label.bind("<Button-3>", open_settings)  # Bind right-click event
clock_label.pack(expand=True)

# Menu
toolbar = ttk.Frame(root)
toolbar.pack(side=tk.TOP, padx=10)

style = ttk.Style()
style.configure("Custom.TMenubutton", background='white', foreground='black')

options = ["Prepared Speech", "Table Topics", "Ice Breaker", "Countdown", "Custom Timer"]
selected_option = tk.StringVar()
selected_option.set(options[0])

menu = ttk.OptionMenu(toolbar, selected_option, *options, command=set_option, style="Custom.TMenubutton")
menu.pack(side=tk.TOP, padx=10)

# Entry Frame for Data Input
entry_frame = tk.Frame(root)
entry_frame.pack(side=tk.TOP, padx=10, pady=10)

date_label = tk.Label(entry_frame, text="Date:", font=("Arial", FONT_SIZE), bg=BACKGROUND_COLOR, fg='black')
date_label.pack(side=tk.LEFT)

date_entry = tk.Entry(entry_frame, font=("Arial", FONT_SIZE))
date_entry.pack(side=tk.LEFT, padx=10)

name_label = tk.Label(entry_frame, text="Member Name:", font=("Arial", FONT_SIZE), bg=BACKGROUND_COLOR, fg='black')
name_label.pack(side=tk.LEFT, padx=10)

name_entry = tk.Entry(entry_frame, font=("Arial", FONT_SIZE))
name_entry.pack(side=tk.LEFT, padx=10)

save_button = tk.Button(entry_frame, text="Save", command=save_data)
save_button.pack(side=tk.LEFT, padx=10)

# Data Storage
members_data = []

# Start the clock
update_background()
update_clock()

root.mainloop()
package com.example.timer
